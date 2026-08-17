from __future__ import annotations

import csv
import json
import uuid
from pathlib import Path
from typing import Any

from .config import settings

_SUPPORTED_SUFFIXES = {".txt", ".md", ".markdown", ".docx", ".pdf", ".csv", ".json", ".log"}


def parse_tabular(file_path: Path) -> list[dict[str, Any]]:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        with file_path.open("r", encoding="utf-8-sig", errors="replace") as file:
            return [dict(row) for row in csv.DictReader(file)]
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("批量导入 Excel 需要安装 openpyxl") from exc
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        if not rows:
            return []
        headers = [str(header or "").strip() for header in rows[0]]
        result: list[dict[str, Any]] = []
        for values in rows[1:]:
            row = {headers[index]: ("" if index >= len(values) or values[index] is None else values[index]) for index in range(len(headers))}
            if any(str(value).strip() for value in row.values()):
                result.append(row)
        return result
    raise ValueError("仅支持 CSV 或 Excel 文件")


def _row_document_name(row: dict[str, Any], index: int) -> str:
    raw = str(row.get("filename") or row.get("name") or row.get("标题") or row.get("title") or f"row_{index + 1}.txt")
    if not Path(raw).suffix:
        raw = f"{raw}.txt"
    return Path(raw).name


def _row_content(row: dict[str, Any]) -> str:
    content = str(row.get("content") or row.get("text") or row.get("正文") or row.get("内容") or "")
    if content.strip():
        return content
    return "\n".join(f"{key}: {value}" for key, value in row.items() if str(value).strip())


def import_tabular(service: Any, actor: Any, kb_id: str, file_path: Path, tags: list[str] | None = None, mode: str = "document") -> dict[str, Any]:
    rows = parse_tabular(file_path)
    imported = 0
    skipped = 0
    errors: list[str] = []
    temp_dir = settings.data_dir / "batch" / kb_id
    temp_dir.mkdir(parents=True, exist_ok=True)
    for index, row in enumerate(rows):
        name = _row_document_name(row, index)
        content = _row_content(row)
        if not content.strip():
            continue
        temp = temp_dir / f"{uuid.uuid4().hex[:8]}_{name}"
        temp.write_text(content, encoding="utf-8")
        try:
            service.ingest_path_with_mode(actor, kb_id, temp, tags, "public")
            imported += 1
        except FileExistsError:
            skipped += 1
        except Exception as exc:
            errors.append(f"{name}: {exc}")
        finally:
            temp.unlink(missing_ok=True)
    return {"imported": imported, "skipped": skipped, "errors": errors[:20], "mode": mode}


def index_folder(service: Any, actor: Any, kb_id: str, folder_path: Path, tags: list[str] | None = None) -> dict[str, Any]:
    if not folder_path.exists() or not folder_path.is_dir():
        raise ValueError("文件夹不存在或不是目录")
    imported = 0
    skipped = 0
    errors: list[str] = []
    for path in sorted(folder_path.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in _SUPPORTED_SUFFIXES:
            continue
        try:
            service.ingest_path_with_mode(actor, kb_id, path, tags, "public")
            imported += 1
        except FileExistsError:
            skipped += 1
        except Exception as exc:
            errors.append(f"{path.name}: {exc}")
    return {"imported": imported, "skipped": skipped, "errors": errors[:20], "folder": str(folder_path)}
